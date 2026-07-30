#!/usr/bin/env python3
"""
build_catalog.py - regenerate the MCR field catalog from an NMLS batch XSD.

Turns MCRBatchFileSchemaV{N}.xsd + the official MCR_FV{N}Sample.xlsx into
02_field_catalog_full.sql (mcr.FieldCatalog, mcr.FieldCatalogElement,
mcr.ListCatalog, mcr.ListElementCatalog, staging tables, ValidationResults).

Usage:
    python build_catalog.py MCRBatchFileSchemaV7.xsd MCR_FV7Sample.xlsx \
        -o 02_field_catalog_full.sql

    # after generating for two versions, diff the models:
    python build_catalog.py --diff model_v7.json model_v8.json

Requires: openpyxl  (pip install openpyxl)

Assumptions this script makes about the schema (verify on a new version -
Step 2 of the Upgrade SOP):
  - Root element is <Mcr>; per-state wrappers <Rmla>/<Sssf> carry stateCode.
  - Leaf elements are named ITEMCODE_n where n is the column suffix
    (pairs (1,2)(3,4)(5,6) = amount/count per form column group).
  - Repeating lists are complex items containing <ItemId> plus leaf elements,
    nested under ListSectionOf.../DetailItemList.
  - The sample workbook carries item codes in the leftmost columns per sheet
    and calc/cross-foot notes in the cells to the right.
If a new form version breaks one of these, this script needs a matching
change - the diff step will make that obvious.
"""
import argparse, json, re, sys
import xml.etree.ElementTree as ET

NS = {'xs': 'http://www.w3.org/2001/XMLSchema'}
CODE = re.compile(r'[A-Z]{1,5}[0-9]{2,4}[A-Z]{0,2}[0-9]?')
NOTE_LABEL = {'ACNOTE': 'Explanatory Notes (state RMLA)',
              'FCNOTE': 'Explanatory Notes (Financial Condition)',
              'NOTE':   'Explanatory Notes (company-level RMLA)'}


# ---------------------------------------------------------------- XSD parse
def walk(el, path):
    out = []
    for child in el:
        tag = child.tag.split('}')[1]
        if tag == 'element':
            nm, ty = child.get('name'), child.get('type')
            ct = child.find('xs:complexType', NS)
            if ct is not None:
                out.append((path, nm, 'COMPLEX'))
                out += walk(ct, path + '/' + nm)
            else:
                out.append((path, nm, ty))
        elif tag in ('sequence', 'all', 'choice', 'complexType'):
            out += walk(child, path)
    return out


def parse_xsd(xsd_path):
    root = ET.parse(xsd_path).getroot()
    mcr = [e for e in root.findall('xs:element', NS) if e.get('name') == 'Mcr'][0]
    rows = walk(mcr.find('xs:complexType', NS), 'Mcr')

    # repeating lists: complex nodes whose children include ItemId
    complexes = [(p, n) for p, n, t in rows if t == 'COMPLEX']
    children = {}
    for p, n, t in rows:
        children.setdefault(p, []).append((n, t))
    list_names = {n for p, n in complexes
                  if any(c == 'ItemId' for c, _ in children.get(p + '/' + n, []))}

    def stem(e):
        if '_' in e:
            c, s = e.rsplit('_', 1)
            if s.isdigit():
                return c, int(s)
        return e, 0

    items, lists, order = {}, {}, 0
    for p, n, t in rows:
        if t == 'COMPLEX':
            continue
        parent = p.rsplit('/', 1)[-1]
        if parent in list_names:
            lists.setdefault(parent, []).append((n, t))
            continue
        order += 1
        code, suf = stem(n)
        if code in ('ACNOTE', 'FCNOTE', 'NOTE'):
            code = n.split('_')[0]
        it = items.setdefault(code, {'path': p, 'scope': scope_of(p), 'els': []})
        it['els'].append((n, t.replace('xs:', ''), order, suf))

    # list parent paths + maxOccurs
    xsd_text = open(xsd_path).read()
    listmeta = {}
    for ln in lists:
        m = re.search(rf'name="{ln}"[^>]*maxOccurs="(\d+|unbounded)"', xsd_text)
        mx = 10000 if (not m or m.group(1) == 'unbounded') else int(m.group(1))
        pp = next(p for p, n, t in rows if n == ln and t == 'COMPLEX')
        listmeta[ln] = {'parent': pp, 'scope': scope_of(pp), 'max': mx}
    return items, lists, listmeta


def scope_of(p):
    if p.startswith('Mcr/Rmlag'):
        return 'COMPANY'
    if p.startswith('Mcr/Rmla'):
        return 'STATE'
    if p.startswith('Mcr/Fc'):
        return 'FC'
    if p.startswith('Mcr/Sssf'):
        return 'SSSF'
    return '?'


# ------------------------------------------------------------ label harvest
DEFAULT_PLAN = "RMLA:0;Company-Level RMLA:0,1,2;FC:2,1,0;SSSF:0,1"


def harvest_labels(xlsx_path, plan_spec=DEFAULT_PLAN):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    # plan_spec: "Sheet:col,col;Sheet2:col" - candidate code columns per sheet
    plans = [(p.split(':')[0], [int(c) for c in p.split(':')[1].split(',')])
             for p in plan_spec.split(';')]
    meta, seq = {}, 0
    for sheet, cols in plans:
        if sheet not in wb.sheetnames:
            print(f'  WARNING: sheet "{sheet}" not in workbook - labels missing for it',
                  file=sys.stderr)
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(values_only=True):
            for ci in cols:
                v = r[ci] if ci < len(r) else None
                if v is None:
                    continue
                code = str(v).strip()
                if not CODE.fullmatch(code):
                    continue
                label = None
                for li in range(ci + 1, min(ci + 3, len(r))):
                    lv = r[li]
                    if lv and not CODE.fullmatch(str(lv).strip()):
                        label = str(lv).strip()
                        break
                rest = ' '.join(str(x) for x in r[ci + 2:ci + 12] if x is not None)
                rest = dedup_tokens(re.sub(r'=?\(?\s*Amount\s*/\s*Count\s*\)?', '',
                                           rest.replace('\xa0', ' ')))
                if code not in meta:
                    seq += 1
                    meta[code] = {'label': (label or code)[:240], 'sheet': sheet,
                                  'note': rest.strip(' =')[:280], 'order': seq}
                break
    return meta


def dedup_tokens(s):
    toks = re.sub(r'\s+', ' ', s).strip().split(' ')
    n = len(toks)
    for plen in range(1, n // 2 + 1):
        if n % plen:
            continue
        seg = toks[:plen]
        if all(toks[i * plen:(i + 1) * plen] == seg for i in range(n // plen)):
            return ' '.join(seg)
    if n >= 8:
        head = ' '.join(toks[:4])
        j = ' '.join(toks).find(head, len(head))
        if j > 0:
            return ' '.join(toks)[:j].strip(' .;')
    return ' '.join(toks)


# ------------------------------------------------------------------ emit SQL
def q(s):
    return (s or '').replace("'", "''").replace('\n', ' ').strip()


def batched(table, cols, rows, fmt, batch=500):
    out = []
    for i in range(0, len(rows), batch):
        out.append(f'INSERT INTO {table} ({cols}) VALUES\n'
                   + ',\n'.join(fmt(r) for r in rows[i:i + batch]) + ';\n')
    return '\nGO\n'.join(out)


DDL = """/* GENERATED by build_catalog.py - do not hand-edit; regenerate instead */
USE MCR_Toolkit;
GO
IF OBJECT_ID('mcr.RepeatingValues')     IS NOT NULL DROP TABLE mcr.RepeatingValues;
IF OBJECT_ID('mcr.ReportValues')        IS NOT NULL DROP TABLE mcr.ReportValues;
IF OBJECT_ID('mcr.ListElementCatalog')  IS NOT NULL DROP TABLE mcr.ListElementCatalog;
IF OBJECT_ID('mcr.ListCatalog')         IS NOT NULL DROP TABLE mcr.ListCatalog;
IF OBJECT_ID('mcr.FieldCatalogElement') IS NOT NULL DROP TABLE mcr.FieldCatalogElement;
IF OBJECT_ID('mcr.FieldCatalog')        IS NOT NULL DROP TABLE mcr.FieldCatalog;
IF OBJECT_ID('mcr.ValidationResults')   IS NOT NULL DROP TABLE mcr.ValidationResults;
GO
CREATE TABLE mcr.FieldCatalog (
    ItemCode      VARCHAR(20)  NOT NULL PRIMARY KEY,
    SectionPath   VARCHAR(120) NOT NULL,
    Scope         VARCHAR(10)  NOT NULL,
    IsCalculated  BIT          NOT NULL,
    IsRequired    BIT          NOT NULL,
    CrossFootNote VARCHAR(280) NULL,
    FormOrder     INT          NOT NULL,
    Label         VARCHAR(240) NOT NULL,
    CONSTRAINT CK_FC_Scope CHECK (Scope IN ('STATE','COMPANY','FC','SSSF'))
);
CREATE TABLE mcr.FieldCatalogElement (
    ElementName VARCHAR(20) NOT NULL PRIMARY KEY,
    ItemCode    VARCHAR(20) NOT NULL REFERENCES mcr.FieldCatalog(ItemCode),
    DataType    VARCHAR(20) NOT NULL,
    ColumnNo    TINYINT     NOT NULL,
    ElemOrder   INT         NOT NULL
);
CREATE TABLE mcr.ListCatalog (
    ListName    VARCHAR(60)  NOT NULL PRIMARY KEY,
    ParentPath  VARCHAR(120) NOT NULL,
    ItemElement VARCHAR(60)  NOT NULL,
    Scope       VARCHAR(10)  NOT NULL,
    MaxItems    INT          NOT NULL
);
CREATE TABLE mcr.ListElementCatalog (
    ListName    VARCHAR(60) NOT NULL REFERENCES mcr.ListCatalog(ListName),
    ElementName VARCHAR(20) NOT NULL,
    DataType    VARCHAR(20) NOT NULL,
    ElemOrder   INT         NOT NULL,
    CONSTRAINT PK_LEC PRIMARY KEY (ListName, ElementName)
);
CREATE TABLE mcr.ReportValues (
    FilingId    INT           NOT NULL,
    ScopeKey    VARCHAR(10)   NOT NULL,
    ElementName VARCHAR(20)   NOT NULL REFERENCES mcr.FieldCatalogElement(ElementName),
    NumValue    DECIMAL(15,2) NULL,
    TextValue   NVARCHAR(4000) NULL,
    CONSTRAINT PK_ReportValues PRIMARY KEY (FilingId, ScopeKey, ElementName)
);
CREATE TABLE mcr.RepeatingValues (
    FilingId    INT           NOT NULL,
    ScopeKey    VARCHAR(10)   NOT NULL,
    ListName    VARCHAR(60)   NOT NULL REFERENCES mcr.ListCatalog(ListName),
    ItemSeq     INT           NOT NULL,
    ElementName VARCHAR(20)   NOT NULL,
    NumValue    DECIMAL(15,2) NULL,
    TextValue   NVARCHAR(400) NULL,
    CONSTRAINT PK_RepeatingValues PRIMARY KEY (FilingId, ScopeKey, ListName, ItemSeq, ElementName)
);
CREATE TABLE mcr.ValidationResults (
    ResultId  INT IDENTITY(1,1) PRIMARY KEY,
    FilingId  INT          NOT NULL,
    Severity  VARCHAR(10)  NOT NULL,
    RuleType  VARCHAR(30)  NOT NULL,
    ScopeKey  VARCHAR(10)  NULL,
    ItemCode  VARCHAR(20)  NULL,
    Detail    VARCHAR(400) NOT NULL,
    CheckedAt DATETIME2    NOT NULL DEFAULT SYSUTCDATETIME()
);
GO
"""

CALC_SCOPE = {'RMLA': 'STATE', 'Company-Level RMLA': 'COMPANY',
              'FC': 'FC', 'SSSF': 'SSSF'}
REQUIRED = ('AC010', 'AC020', 'AC070')


def emit(items, lists, listmeta, labels, out_path, model_path):
    cat, els = [], []
    for code, it in items.items():
        lab = labels.get(code, {})
        label = NOTE_LABEL.get(code) or lab.get('label') or code
        cat.append((code, it['path'], it['scope'], 0,
                    1 if code in REQUIRED else 0,
                    q(lab.get('note', ''))[:280], lab.get('order', 100000),
                    q(label)[:240]))
        for name, typ, eorder, suf in it['els']:
            els.append((name, code, typ, suf, eorder))
    calconly = sorted(c for c in labels if c not in items)
    for code in calconly:
        lab = labels[code]
        cat.append((code, '', CALC_SCOPE.get(lab['sheet'], 'STATE'), 1, 0,
                    q(lab.get('note', ''))[:280], lab['order'],
                    q(lab['label'])[:240]))
    cat.sort(key=lambda r: (r[6], r[0]))

    listcat, listels = [], []
    for ln, meta in listmeta.items():
        listcat.append((ln, meta['parent'], ln, meta['scope'], meta['max']))
        o = 0
        for n, t in lists[ln]:
            if n == 'ItemId':
                continue
            o += 1
            listels.append((ln, n, t.replace('xs:', ''), o))

    parts = [DDL,
             batched('mcr.FieldCatalog',
                     'ItemCode,SectionPath,Scope,IsCalculated,IsRequired,CrossFootNote,FormOrder,Label',
                     cat,
                     lambda r: f"('{r[0]}','{r[1]}','{r[2]}',{r[3]},{r[4]},"
                               + (f"'{r[5]}'" if r[5] else 'NULL')
                               + f",{r[6]},'{r[7]}')"),
             '\nGO\n',
             batched('mcr.FieldCatalogElement',
                     'ElementName,ItemCode,DataType,ColumnNo,ElemOrder',
                     sorted(els, key=lambda e: e[4]),
                     lambda r: f"('{r[0]}','{r[1]}','{r[2]}',{r[3]},{r[4]})"),
             '\nGO\n',
             batched('mcr.ListCatalog', 'ListName,ParentPath,ItemElement,Scope,MaxItems',
                     listcat, lambda r: f"('{r[0]}','{r[1]}','{r[2]}','{r[3]}',{r[4]})"),
             '\nGO\n',
             batched('mcr.ListElementCatalog', 'ListName,ElementName,DataType,ElemOrder',
                     listels, lambda r: f"('{r[0]}','{r[1]}','{r[2]}',{r[3]})"),
             f"""
GO
DECLARE @i INT, @e INT;
SELECT @i=COUNT(*) FROM mcr.FieldCatalog; SELECT @e=COUNT(*) FROM mcr.FieldCatalogElement;
PRINT '02 complete: ' + CAST(@i AS VARCHAR(10)) + ' items, ' + CAST(@e AS VARCHAR(10)) + ' elements cataloged.';
GO
"""]
    open(out_path, 'w').write('\n'.join(parts))
    json.dump({'items': items, 'lists': lists, 'calconly': calconly},
              open(model_path, 'w'))
    print(f'{out_path}: {len(cat)} items ({len(calconly)} calculated), '
          f'{len(els)} elements, {len(listels)} list elements -> {model_path}')


# ---------------------------------------------------------------------- diff
def diff(old_path, new_path):
    o, n = json.load(open(old_path)), json.load(open(new_path))
    oi, ni = set(o['items']), set(n['items'])
    print('ITEMS  added:', sorted(ni - oi) or '-')
    print('ITEMS  removed:', sorted(oi - ni) or '-')
    oc, nc = set(o['calconly']), set(n['calconly'])
    print('CALC   added:', sorted(nc - oc) or '-')
    print('CALC   removed:', sorted(oc - nc) or '-')
    oe = {e[0]: (c, e[1]) for c, it in o['items'].items() for e in it['els']}
    ne = {e[0]: (c, e[1]) for c, it in n['items'].items() for e in it['els']}
    print('ELEMENTS added:', len(set(ne) - set(oe)), 'removed:', len(set(oe) - set(ne)))
    retyped = [k for k in set(oe) & set(ne) if oe[k][1] != ne[k][1]]
    print('ELEMENTS retyped:', retyped or '-')
    moved = [c for c in oi & ni if o['items'][c]['path'] != n['items'][c]['path']]
    print('ITEMS  moved section:', moved or '-')
    ol, nl = set(o['lists']), set(n['lists'])
    print('LISTS  added:', sorted(nl - ol) or '-', ' removed:', sorted(ol - nl) or '-')
    print('\nReview every line above against 03 (loader mappings), 04 (targeted')
    print('rules), and 05 (assembly) per the Upgrade SOP.')


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('xsd', nargs='?')
    ap.add_argument('workbook', nargs='?')
    ap.add_argument('-o', '--out', default='02_field_catalog_full.sql')
    ap.add_argument('-m', '--model', default='model.json')
    ap.add_argument('--sheets', default=None,
                    help='harvest plan "Sheet:col,col;Sheet2:col" (default: FV7 layout)')
    ap.add_argument('--diff', nargs=2, metavar=('OLD_MODEL', 'NEW_MODEL'))
    a = ap.parse_args()
    if a.diff:
        diff(*a.diff)
    elif a.xsd and a.workbook:
        items, lists, listmeta = parse_xsd(a.xsd)
        labels = harvest_labels(a.workbook, a.sheets or DEFAULT_PLAN)
        missing = [c for c in items if c not in labels
                   and not c.endswith('NOTE')]
        if missing:
            print(f'  note: {len(missing)} XSD items have no workbook label '
                  f'(labeled with their code): {missing[:10]}', file=sys.stderr)
        emit(items, lists, listmeta, labels, a.out, a.model)
    else:
        ap.print_help()
